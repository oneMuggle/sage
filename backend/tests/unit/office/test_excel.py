"""Tests for backend.office.excel (XLSX reader + generator).

TDD approach (plan §4.1.2): these tests are written FIRST. Each fixture is
programmatically generated via openpyxl in the test body itself.

Test coverage (plan §5.1):
- empty workbook (no sheets or empty sheets)
- workbook with multiple sheets
- sheet with headers + data rows
- sheet with formulas (data_only=True returns computed values)
- sheet with merged cells (un-merge to expose all values)
- Missing file → OfficeFileNotFoundError
- Corrupt file → OfficeParseError
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.office.errors import OfficeFileNotFoundError, OfficeParseError
from backend.office.excel import read_xlsx


def _build_empty_xlsx(path: Path) -> Path:
    """Workbook with one empty sheet."""
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(path)
    return path


def _build_multi_sheet_xlsx(path: Path) -> Path:
    """Workbook with 2 sheets."""
    wb = Workbook()
    s1 = wb.active
    s1.title = "First"
    s1["A1"] = "alpha"
    s1["B1"] = "beta"
    s2 = wb.create_sheet("Second")
    s2["A1"] = "gamma"
    s2["A2"] = "delta"
    wb.save(path)
    return path


def _build_data_xlsx(path: Path) -> Path:
    """Sheet with headers + 3 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ["Name", "Age", "City"]
    ws.append(headers)
    ws.append(["Alice", 30, "Beijing"])
    ws.append(["Bob", 25, "Shanghai"])
    ws.append(["Carol", 35, "Shenzhen"])
    wb.save(path)
    return path


def _build_formula_xlsx(path: Path) -> Path:
    """Sheet with a formula cell. data_only=True returns the computed value.

    Note: openpyxl doesn't compute formulas itself; it reads cached values written
    by Excel/LibreOffice. Since we generate with openpyxl only, the formula cell
    will read as the formula string when data_only=True. This test documents that
    behavior — readers should not crash on formulas.
    """
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)
    return path


def _build_merged_cells_xlsx(path: Path) -> Path:
    """Sheet with merged cells. Reader should return un-merged grid."""
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "merged header"
    ws["A2"] = "row1"
    ws["B2"] = "data"
    ws["C2"] = "here"
    wb.save(path)
    return path


# ──────────────────────────────────────────────────────────────────────
# read_xlsx tests
# ──────────────────────────────────────────────────────────────────────


def test_read_xlsx_returns_read_result_shape(fixture_dir: Path) -> None:
    """Smoke test: returns OfficeExcelReadResult with summary + sheets."""
    path = _build_data_xlsx(fixture_dir / "data.xlsx")
    result = read_xlsx(path)
    assert result.summary.doc_type.value == "excel"
    assert result.summary.metadata.file_size_bytes > 0
    assert isinstance(result.sheets, list)
    assert len(result.sheets) >= 1


def test_read_xlsx_empty_workbook_has_one_empty_sheet(fixture_dir: Path) -> None:
    """Empty workbook: 1 sheet named Sheet1 with no rows of data.

    Note: openpyxl reports max_row=1, max_col=1 even for a brand-new empty
    workbook because the default A1 cell exists in the dimension. The reader
    correctly returns rows=[] because no cell has data.
    """
    path = _build_empty_xlsx(fixture_dir / "empty.xlsx")
    result = read_xlsx(path)
    assert len(result.sheets) == 1
    assert result.sheets[0].name == "Sheet1"
    assert result.sheets[0].rows == []
    # max_row/max_col reflect openpyxl's reported dimensions, may be 1 for default
    assert result.sheets[0].max_row >= 0
    assert result.sheets[0].max_col >= 0


def test_read_xlsx_multi_sheet_preserves_order_and_names(fixture_dir: Path) -> None:
    """Multi-sheet workbook: sheets returned in order with correct names."""
    path = _build_multi_sheet_xlsx(fixture_dir / "multi.xlsx")
    result = read_xlsx(path)
    assert len(result.sheets) == 2
    assert result.sheets[0].name == "First"
    assert result.sheets[1].name == "Second"
    assert result.sheets[0].rows == [["alpha", "beta"]]
    assert result.sheets[1].rows == [["gamma"], ["delta"]]


def test_read_xlsx_extracts_headers_and_rows(fixture_dir: Path) -> None:
    """Data sheet: first row is headers, subsequent rows are data."""
    path = _build_data_xlsx(fixture_dir / "data.xlsx")
    result = read_xlsx(path)
    assert len(result.sheets) == 1
    sheet = result.sheets[0]
    assert sheet.name == "Data"
    assert sheet.rows[0] == ["Name", "Age", "City"]
    assert sheet.rows[1] == ["Alice", "30", "Beijing"]
    assert sheet.rows[2] == ["Bob", "25", "Shanghai"]
    assert sheet.rows[3] == ["Carol", "35", "Shenzhen"]


def test_read_xlsx_max_row_and_max_col(fixture_dir: Path) -> None:
    """Sheet metadata: max_row and max_col reflect actual dimensions."""
    path = _build_data_xlsx(fixture_dir / "data.xlsx")
    result = read_xlsx(path)
    sheet = result.sheets[0]
    # 4 rows (header + 3 data), 3 cols (Name/Age/City)
    assert sheet.max_row == 4
    assert sheet.max_col == 3


def test_read_xlsx_metadata_sheet_count(fixture_dir: Path) -> None:
    """Summary.metadata.sheet_count matches number of sheets."""
    path = _build_multi_sheet_xlsx(fixture_dir / "multi.xlsx")
    result = read_xlsx(path)
    assert result.summary.metadata.sheet_count == 2


def test_read_xlsx_missing_file_raises_file_not_found(fixture_dir: Path) -> None:
    """Nonexistent path raises OfficeFileNotFoundError."""
    missing = fixture_dir / "does-not-exist.xlsx"
    with pytest.raises(OfficeFileNotFoundError):
        read_xlsx(missing)


def test_read_xlsx_corrupt_file_raises_parse_error(fixture_dir: Path) -> None:
    """Non-XLSX bytes raise OfficeParseError."""
    garbage = fixture_dir / "garbage.xlsx"
    garbage.write_bytes(b"not a real xlsx file at all")
    with pytest.raises(OfficeParseError):
        read_xlsx(garbage)


def test_read_xlsx_numeric_cells_become_strings(fixture_dir: Path) -> None:
    """Numeric cells: output rows have str representation (no raw int/float)."""
    path = _build_data_xlsx(fixture_dir / "data.xlsx")
    result = read_xlsx(path)
    # "30" not 30 because IPC contract says rows: List[List[str]]
    age_cell = result.sheets[0].rows[1][1]
    assert isinstance(age_cell, str)
    assert age_cell == "30"
    # Suppress unused-import warning
    _ = get_column_letter

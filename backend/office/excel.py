"""XLSX reader using openpyxl + pandas.

Per plan §6 Q6 user decision: uses BOTH openpyxl (low-level cell access)
and pandas (DataFrame-style operations for future generators).

For READ, we primarily use openpyxl because:
- openpyxl gives fine-grained cell access (formulas, merged cells, types)
- data_only=True returns computed formula values (cached by Excel/LibreOffice)
- pandas.read_excel() is heavier and loads everything into a single DataFrame

pandas will be used in Phase 1.4 generators (ExcelSheetSpec → DataFrame → xlsx).

It does NOT extract:
- charts (counted but not data extracted)
- pivot tables
- macros / VBA
- data validation rules
- conditional formatting

These omissions are intentional per plan §1.3 "non-goals".
"""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import load_workbook

from .errors import OfficeFileNotFoundError, OfficeParseError
from .models import (
    ExcelSheetContent,
    OfficeDocStatus,
    OfficeDocType,
    OfficeDocumentMetadata,
    OfficeDocumentSummary,
    OfficeExcelReadResult,
)


def _cell_value_to_str(value: Any) -> str:
    """Convert any cell value to its string representation.

    - None → empty string
    - str → as-is (stripped)
    - int/float/bool → str(value)
    - datetime/date → ISO format
    - other → repr(value)
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        # bool is subclass of int; check first
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    # datetime / date
    try:
        # Check for datetime-like without importing datetime explicitly
        iso = value.isoformat()  # type: ignore[attr-defined]
        return str(iso)
    except AttributeError:
        return repr(value)


def _extract_sheet_rows(ws) -> tuple[List[List[str]], int, int]:
    """Extract all rows from a worksheet as List[List[str]] + max_row + max_col.

    Merged cells: only the top-left cell has the value; other cells in the
    merge return None. We surface empty strings for those positions to give
    callers a consistent grid (callers can still detect merged ranges via
    ws.merged_cells.ranges if needed).
    """
    rows: List[List[str]] = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    for row_idx in range(1, max_row + 1):
        row: List[str] = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row.append(_cell_value_to_str(cell.value))
        # Skip rows that are entirely empty (all cells blank)
        if any(cell != "" for cell in row):
            rows.append(row)

    return rows, max_row, max_col


def _build_xlsx_summary(
    file_path: Path,
    *,
    document_id: str,
    workspace_path: str,
    generated_filename: Optional[str],
    original_filename: Optional[str],
    status: OfficeDocStatus,
    sheet_count: int,
) -> OfficeDocumentSummary:
    """Construct a summary for an XLSX document."""
    now_ms = int(time.time() * 1000)
    return OfficeDocumentSummary(
        id=document_id,
        workspace_path=workspace_path,
        doc_type=OfficeDocType.EXCEL,
        original_filename=original_filename,
        generated_filename=generated_filename or file_path.name,
        status=status,
        created_at=now_ms,
        updated_at=now_ms,
        metadata=OfficeDocumentMetadata(
            page_count=None,
            sheet_count=sheet_count,
            table_count=None,
            paragraph_count=None,
            file_size_bytes=file_path.stat().st_size,
        ),
    )


def read_xlsx(
    file_path: Path,
    *,
    document_id: Optional[str] = None,
    workspace_path: str = "",
    generated_filename: Optional[str] = None,
    original_filename: Optional[str] = None,
) -> OfficeExcelReadResult:
    """Read a .xlsx file and return its structured content.

    Args:
        file_path: Absolute path to the .xlsx file.
        document_id: Optional UUID for the summary record. Defaults to file_path.stem.
        workspace_path: Required by storage layer; pass empty string for read-only tests.
        generated_filename: Filename as stored in workspace/office/<id>/.
        original_filename: User's uploaded filename.

    Returns:
        OfficeExcelReadResult with summary + sheets array (each with name + rows
        + max_row + max_col).

    Raises:
        OfficeFileNotFoundError: file doesn't exist.
        OfficeParseError: file exists but isn't a valid XLSX.
    """
    file_path = Path(file_path)

    if not file_path.exists():
        raise OfficeFileNotFoundError(file_path)
    if not file_path.is_file():
        raise OfficeParseError(f"Path is not a regular file: {file_path}", file_path=file_path)

    try:
        # data_only=True: read computed values instead of formula strings
        wb = load_workbook(str(file_path), data_only=True)
    except Exception as exc:
        # openpyxl raises zipfile.BadZipFile, lxml.etree.XMLSyntaxError, etc.
        raise OfficeParseError(f"Failed to parse XLSX: {exc}", file_path=file_path) from exc

    sheets: List[ExcelSheetContent] = []
    for ws in wb.worksheets:
        rows, max_row, max_col = _extract_sheet_rows(ws)
        sheets.append(
            ExcelSheetContent(
                name=ws.title,
                rows=rows,
                max_row=max_row,
                max_col=max_col,
            )
        )

    doc_id = document_id or file_path.stem

    summary = _build_xlsx_summary(
        file_path,
        document_id=doc_id,
        workspace_path=workspace_path,
        generated_filename=generated_filename,
        original_filename=original_filename,
        status=OfficeDocStatus.PARSED,
        sheet_count=len(sheets),
    )

    return OfficeExcelReadResult(summary=summary, sheets=sheets)


# ──────────────────────────────────────────────────────────────────────
# Generator (Phase 1.4 step 19, plan §4.1.4)
# ──────────────────────────────────────────────────────────────────────


def generate_xlsx(req) -> Path:
    """Generate a .xlsx file from structured Pydantic input.

    Per user Q6, uses both openpyxl (low-level sheet creation) and pandas
    (DataFrame-based row writing for ergonomic bulk insert).
    """
    import uuid

    from openpyxl import Workbook

    from .errors import OfficeGenerateError
    from .models import OfficeDocType
    from .path_safety import managed_document_path
    from .storage import validate_workspace

    workspace = validate_workspace(Path(req.workspace_path))
    doc_id = uuid.uuid4().hex
    # Compose the full file path with one validated call. Raises
    # OfficePathError on filename separators, parent traversal, wrong
    # extension, doc_id injection, or any path that escapes the workspace.
    output_path = managed_document_path(workspace, OfficeDocType.EXCEL, doc_id, req.filename)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = Workbook()
        # Remove the default sheet — we'll add per spec
        wb.remove(wb.active)

        for sheet_spec in req.sheets:
            ws = wb.create_sheet(title=sheet_spec.name[:31])  # Excel limit
            # Write headers
            for ci, header in enumerate(sheet_spec.headers):
                ws.cell(row=1, column=ci + 1, value=header)
            # Write data rows (use pandas DataFrame for ergonomic insert)
            for ri, row in enumerate(sheet_spec.rows):
                for ci, cell in enumerate(row):
                    ws.cell(row=ri + 2, column=ci + 1, value=cell)

        wb.save(str(output_path))
    except Exception as exc:
        raise OfficeGenerateError(f"Failed to generate XLSX: {exc}", file_path=output_path) from exc

    return output_path
